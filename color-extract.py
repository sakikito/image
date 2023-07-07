import matplotlib.pyplot as plt
import cv2
import numpy as np

img_num = int(input("画像数: "))

for i in range(1, img_num+1):

    img = cv2.imread(f"image_{i+1}.jpg")
    img_array = np.asarray(img)
    print(len(img_array))

    B = []
    G = []
    R = []

    for x in range(0, len(img_array)):
        for y in range(0, len(img_array)):
            B.append(img_array[x][y][0])
            R.append(img_array[x][y][1])
            R.append(img_array[x][y][2])

    tmp_num = len(img_array)
    
    import openpyxl
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = f"mentaiko{i+1}"

    sheet["A1"].value = f"海苔{i+1}"
    sheet["A2"].value = 'B'
    sheet["B2"].value = 'G'
    sheet["C2"].value = 'R'
    
    for j in range(2, tmp_num+2):
        sheet.cell(column=1, row=j+1, value=B[j-1])
        sheet.cell(column=2, row=j+1, value=G[j-1])
        sheet.cell(column=3, row=j+1, value=R[j-1])

    wb.save(f"nori{i+1}_lab_excel.xlsx")
    wb.close()
