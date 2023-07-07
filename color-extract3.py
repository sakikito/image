import matplotlib.pyplot as plt
import cv2
import numpy as np

img_num = int(input("画像数: ")) + 1

for i in range(1, img_num):

    img = cv2.imread("lab_image_{}.jpg".format(i))
    img_array = np.asarray(img)
    print(len(img_array))

    L = []
    A = []
    B = []

    for x in range(0, len(img_array)):
        for y in range(0, len(img_array)):
            L.append(img_array[x][y][0])
            A.append(img_array[x][y][1])
            B.append(img_array[x][y][2])

    tmp_num = len(img_array)
    
    import openpyxl
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'mentaiko{}'.format(i) 

    sheet["A1"].value = '海苔'.format(i)
    sheet["A2"].value = 'L'
    sheet["B2"].value = 'A'
    sheet["C2"].value = 'B'
    
    for j in range(2, tmp_num+2):
        sheet.cell(column=1, row=j+1, value=L[j-1])
        sheet.cell(column=2, row=j+1, value=A[j-1])
        sheet.cell(column=3, row=j+1, value=B[j-1])

    wb.save('nori{}_lab_excel.xlsx'.format(i))
    wb.close()